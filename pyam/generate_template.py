#!/usr/bin/env python3
# Copyright 2023, Dr John A.R. Williams
# SPDX-License-Identifier: GPL-3.0-only
"""Generate template marking spreadsheet

Starts from from template-template.xlsx
and adds in a row per test with the description, a named cell to filled in
as PASSED or FAILED from the students reports and a mark as per the test
manifest
"""
import argparse
from pathlib import Path
import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname, absolute_coordinate
from pyam.config import CONFIG
from pyam.args import add_common_args
from pyam.cohort import get_cohort


def main(args=None):
    """Main routing - generates a cohort template from template-template"""
    if args is None:
        parser = argparse.ArgumentParser(description=__doc__)
        add_args(parser)
        args = parser.parse_args()
    cohort = get_cohort(args.cohort)
    destination = cohort.report_path / "template.xlsx"
    if args.output:
        destination = args.output
    if not (args.overwrite) and destination.exists():
        raise FileExistsError(destination)
    cohort.start_log_section(
        f"Generate template {destination.relative_to(CONFIG.root_path)}")
    template = openpyxl.load_workbook(
        filename=str(Path(__file__).parent / "template-template.xlsx"))
    worksheet = template.worksheets[0]
    worksheet["A2"].value = cohort.get("module_code")
    worksheet["A3"].value = cohort.get("assessment")
    row = args.start_row
    for test, details in cohort.tests().items():
        bcell = worksheet.cell(row=row, column=2)
        ccell = worksheet.cell(row=row, column=3)
        dcell = worksheet.cell(row=row, column=4)
        ref = f"{quote_sheetname(worksheet.title)}!{absolute_coordinate(f'C{row}')}"
        defn = DefinedName(name=test, attr_text=ref)
        template.defined_names.append(defn)
        bcell.value = details.get("description", test)
        ccell.value = "FAILED"
        dcell.value = details.get("mark", 1)
        row += 1
    template.save(destination)


def add_args(parser=argparse.ArgumentParser(description=__doc__)):
    """Return args for this script"""
    add_common_args(parser)
    parser.add_argument('-o',
                        '--output',
                        type=Path,
                        help="Destination path for template spreadsheet")
    parser.add_argument("--start-row",
                        type=int,
                        default=14,
                        help="Row to start filling up fields from")


if __name__ == "__main__":
    main()
